"""
Excel 管理模块 - 提供 Excel 文件的创建、读取、写入和操作功能

主要功能：
- ExcelManager: 轻量级工作簿操作（基于 openpyxl）
- ExcelHandler: 面向已有文件的读取/写入工具
- OpenExcel: 通过 Excel 应用打开工作簿（支持数据刷新）
- ExcelOperation: 数据处理类（拆分、合并等）
"""

from pathlib import Path
from typing import List, Optional, Sequence, Tuple, Union, Iterator, Dict, Any
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from contextlib import contextmanager
from .stringManager import StringBaba


# ============================================================================
# 工具函数
# ============================================================================

def create_workbook(file_path: Union[str, Path], default_sheet: str = "sheet1") -> None:
    """
    创建一个新的 Excel 工作簿
    
    Args:
        file_path: 文件路径
        default_sheet: 默认工作表名称
        
    Raises:
        ValueError: 当 default_sheet 为空时
        IOError: 当文件创建失败时
    """
    if not default_sheet or not isinstance(default_sheet, str):
        raise ValueError("工作表名称必须是有效的字符串")
    
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        if sheet is not None:
            sheet.title = default_sheet
        else:
            wb.create_sheet(title=default_sheet)
        
        # 确保目录存在
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        wb.close()
    except Exception as e:
        raise IOError(f"创建工作簿失败: {e}") from e


def _auto_range(
    start_row: int, 
    start_col: int, 
    data: Sequence[Sequence], 
    use_explicit: int,
    end_row: int, 
    end_col: int
) -> Tuple[int, int]:
    """
    根据传入数据自动计算写入的结束行列
    
    Args:
        start_row: 起始行号
        start_col: 起始列号
        data: 二维数据
        use_explicit: 0=自动计算, 1=使用显式传入的 end_row/end_col
        end_row: 显式结束行号
        end_col: 显式结束列号
        
    Returns:
        (计算后的结束行号, 计算后的结束列号)
    """
    if use_explicit == 0 and data and len(data) > 0:
        calculated_end_row = len(data) + start_row - 1
        calculated_end_col = len(data[0]) + start_col - 1 if data[0] else start_col
        return calculated_end_row, calculated_end_col
    return end_row, end_col


def _apply_styles(
    worksheet: Worksheet, 
    start_row: int, 
    start_col: int, 
    end_row: int, 
    end_col: int,
    header_style: bool = True
) -> None:
    """
    应用样式到单元格范围
    
    Args:
        worksheet: 工作表对象
        start_row: 起始行号
        start_col: 起始列号
        end_row: 结束行号
        end_col: 结束列号
        header_style: 是否为第一行应用表头样式
    """
    # 定义样式
    font = Font(name="Microsoft YaHei", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="0070C0")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    align = Alignment(vertical="center", horizontal="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用样式到所有单元格
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = header_font if (header_style and row == start_row) else font
            cell.alignment = align
            cell.border = thin_border
            if header_style and row == start_row:
                cell.fill = header_fill
    
    # 自动调整列宽
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(start_row, end_row + 1):
            val = worksheet.cell(row=row, column=col).value
            length = len(str(val)) if val is not None else 0
            if length > max_len:
                max_len = length
        worksheet.column_dimensions[letter].width = max(8, min(int(max_len * 1.2) + 2, 50))


# ============================================================================
# 核心类: ExcelManager
# ============================================================================

class ExcelManager:
    """
    轻量级 Excel 工作簿管理类（基于 openpyxl）
    
    功能特性：
    - 自动创建不存在的文件
    - 支持多工作表操作
    - 快速读写数据
    - 自动应用样式
    
    示例:
        >>> # 创建新文件
        >>> wb = ExcelManager("data.xlsx")
        >>> wb.write_sheet("Sheet1", [["Name", "Age"], ["Alice", 25]], 1, 1)
        >>> wb.save()
        
        >>> # 快速写入（自动计算范围）
        >>> wb.fast_write("Sheet1", [["Bob", 30]], start_row=3, start_col=1)
        
        >>> # 读取数据
        >>> data = wb.read_sheet("Sheet1", 1, 1, 2, 2)
    """
    
    def __init__(self, file_path: Union[str, Path, None], default_sheet: str = "sheet1"):
        """
        初始化 ExcelManager
        
        Args:
            file_path: Excel 文件路径
            default_sheet: 默认工作表名称（文件不存在时创建）
            
        Raises:
            FileNotFoundError: 文件路径无效
            ValueError: file_path 为 None
            IOError: 文件加载失败
        """
        if file_path is None:
            raise ValueError("file_path 不能为 None")
            
        self.file_path = Path(file_path)
        self._workbook: Optional[Workbook] = None
        
        if not self.file_path.parent.exists():
            raise FileNotFoundError(f"目录不存在: {self.file_path.parent}")
        
        try:
            if not self.file_path.exists():
                create_workbook(str(self.file_path), default_sheet)
            
            self._workbook = load_workbook(str(self.file_path))
        except Exception as e:
            raise IOError(f"加载工作簿失败: {e}") from e
    
    def __enter__(self) -> "ExcelManager":
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """上下文管理器退出时自动保存"""
        if exc_type is None:
            self.save()
        self.close()
    
    @property
    def workbook(self) -> Workbook:
        """获取底层 Workbook 对象"""
        if self._workbook is None:
            raise RuntimeError("工作簿已关闭")
        return self._workbook
    
    @property
    def sheet_names(self) -> List[str]:
        """获取所有工作表名称"""
        return self.workbook.sheetnames
    
    def _ensure_sheet(self, sheet_name: str) -> Worksheet:
        """
        确保工作表存在，不存在则创建
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            Worksheet 对象
        """
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
        return self.workbook[sheet_name]
    
    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> Worksheet:
        """
        创建新工作表
        
        Args:
            sheet_name: 工作表名称
            index: 插入位置（可选）
            
        Returns:
            新创建的工作表对象
            
        Raises:
            ValueError: 工作表名称已存在
        """
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 已存在")
        
        return self.workbook.create_sheet(title=sheet_name, index=index)
    
    def delete_sheet(self, sheet_name: str) -> None:
        """
        删除工作表
        
        Args:
            sheet_name: 要删除的工作表名称
            
        Raises:
            ValueError: 工作表不存在或是唯一工作表
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        
        if len(self.workbook.sheetnames) == 1:
            raise ValueError("不能删除唯一的工作表")
        
        sheet = self.workbook[sheet_name]
        self.workbook.remove(sheet)
    
    def write_sheet(
        self, 
        sheet_name: str, 
        data: Sequence[Sequence], 
        start_row: int = 1, 
        start_col: int = 1,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None,
        apply_styles: bool = True,
        header_row: bool = True
    ) -> None:
        """
        写入数据到工作表
        
        Args:
            sheet_name: 工作表名称
            data: 二维数据（列表的列表）
            start_row: 起始行号（1-based）
            start_col: 起始列号（1-based）
            end_row: 结束行号（可选，自动计算）
            end_col: 结束列号（可选，自动计算）
            apply_styles: 是否应用样式
            header_row: 第一行是否为表头（影响样式）
            
        Raises:
            ValueError: 数据为空或维度不匹配
        """
        if not data:
            raise ValueError("数据不能为空")
        
        # 自动计算结束行列
        if end_row is None:
            end_row = start_row + len(data) - 1
        if end_col is None:
            max_cols = max(len(row) for row in data) if data else 0
            end_col = start_col + max_cols - 1
        
        worksheet = self._ensure_sheet(sheet_name)
        
        # 写入数据
        for i, row_data in enumerate(data):
            row_idx = start_row + i
            for j, value in enumerate(row_data):
                col_idx = start_col + j
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 应用样式
        if apply_styles:
            _apply_styles(worksheet, start_row, start_col, end_row, end_col, header_row)
    
    def read_sheet(
        self, 
        sheet_name: str, 
        start_row: int = 1, 
        start_col: int = 1,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None
    ) -> List[List[Any]]:
        """
        从工作表读取数据
        
        Args:
            sheet_name: 工作表名称
            start_row: 起始行号（1-based）
            start_col: 起始列号（1-based）
            end_row: 结束行号（可选，读取到最后有数据的行）
            end_col: 结束列号（可选，读取到最后有数据的列）
            
        Returns:
            二维数据列表
            
        Raises:
            ValueError: 工作表不存在
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        
        worksheet = self.workbook[sheet_name]
        
        # 自动检测结束行列
        if end_row is None:
            end_row = worksheet.max_row
        if end_col is None:
            end_col = worksheet.max_column
        
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(worksheet.cell(row=row, column=col).value)
            data.append(row_data)
        
        return data
    
    def fast_write(
        self, 
        sheet_name: str, 
        data: Sequence[Sequence], 
        start_row: int = 1, 
        start_col: int = 1,
        use_explicit_range: bool = False,
        end_row: int = 0,
        end_col: int = 0,
        header_row: bool = True
    ) -> None:
        """
        快速写入数据（简化版，自动计算范围）
        
        Args:
            sheet_name: 工作表名称
            data: 二维数据
            start_row: 起始行号
            start_col: 起始列号
            use_explicit_range: 是否使用显式范围（False=自动计算）
            end_row: 显式结束行号（use_explicit_range=True 时使用）
            end_col: 显式结束列号（use_explicit_range=True 时使用）
            header_row: 第一行是否为表头
        """
        actual_end_row, actual_end_col = _auto_range(
            start_row, start_col, data, 1 if use_explicit_range else 0, end_row, end_col
        )
        self.write_sheet(
            sheet_name, data, start_row, start_col, 
            actual_end_row, actual_end_col, 
            apply_styles=True, header_row=header_row
        )
    
    def write_dataframe(
        self, 
        sheet_name: str, 
        df: pd.DataFrame, 
        start_row: int = 1, 
        start_col: int = 1,
        include_header: bool = True,
        index: bool = False
    ) -> None:
        """
        写入 pandas DataFrame 到工作表
        
        Args:
            sheet_name: 工作表名称
            df: DataFrame 数据
            start_row: 起始行号
            start_col: 起始列号
            include_header: 是否包含列名
            index: 是否包含行索引
        """
        # 转换 DataFrame 为列表
        data = []
        
        if include_header:
            headers = list(df.columns)
            if index:
                headers = [df.index.name or ''] + headers
            data.append(headers)
        
        for idx, row in df.iterrows():
            row_data = list(row)
            if index:
                row_data = [idx] + row_data
            data.append(row_data)
        
        self.write_sheet(
            sheet_name, data, start_row, start_col,
            apply_styles=True, header_row=include_header
        )
    
    def read_dataframe(
        self, 
        sheet_name: str, 
        start_row: int = 1, 
        header_row: int = 1
    ) -> pd.DataFrame:
        """
        从工作表读取为 pandas DataFrame
        
        Args:
            sheet_name: 工作表名称
            start_row: 数据起始行号
            header_row: 表头所在行号
            
        Returns:
            DataFrame 对象
        """
        data = self.read_sheet(sheet_name, start_row=start_row)
        
        if not data:
            return pd.DataFrame()
        
        # 第一行作为表头
        headers = [str(h) for h in data[0]]
        rows = data[1:] if len(data) > 1 else []
        
        # type: ignore 用于解决 pandas 类型提示问题
        return pd.DataFrame(rows, columns=headers)  # type: ignore
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """
        获取工作表信息
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            包含工作表信息的字典
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        
        worksheet = self.workbook[sheet_name]
        return {
            "name": sheet_name,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "dimensions": worksheet.dimensions,
            "index": self.workbook.sheetnames.index(sheet_name)
        }
    
    def copy_sheet(self, source_name: str, target_name: str) -> Worksheet:
        """
        复制工作表
        
        Args:
            source_name: 源工作表名称
            target_name: 目标工作表名称
            
        Returns:
            新创建的工作表对象
        """
        if source_name not in self.workbook.sheetnames:
            raise ValueError(f"源工作表 '{source_name}' 不存在")
        
        if target_name in self.workbook.sheetnames:
            raise ValueError(f"目标工作表 '{target_name}' 已存在")
        
        source = self.workbook[source_name]
        return self.workbook.copy_worksheet(source)
    
    def save(self, file_path: Optional[Union[str, Path]] = None) -> None:
        """
        保存工作簿
        
        Args:
            file_path: 保存路径（可选，默认为原路径）
            
        Raises:
            IOError: 保存失败
        """
        save_path = file_path or self.file_path
        
        try:
            # 确保目录存在
            Path(save_path).parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(save_path)
        except Exception as e:
            raise IOError(f"保存工作簿失败: {e}") from e
    
    def close(self) -> None:
        """关闭工作簿并释放资源"""
        if self._workbook:
            try:
                self._workbook.close()
            except Exception:
                pass
            finally:
                self._workbook = None
    
    @classmethod
    def create(
        cls, 
        file_path: Union[str, Path], 
        default_sheet: str = "sheet1",
        overwrite: bool = False
    ) -> "ExcelManager":
        """
        类方法：创建新的 ExcelManager 实例
        
        Args:
            file_path: 文件路径
            default_sheet: 默认工作表名称
            overwrite: 是否覆盖已存在的文件
            
        Returns:
            ExcelManager 实例
            
        Raises:
            FileExistsError: 文件已存在且 overwrite=False
        """
        path = Path(file_path)
        
        if path.exists() and not overwrite:
            raise FileExistsError(f"文件已存在: {path}")
        
        if path.exists():
            path.unlink()
        
        return cls(file_path, default_sheet)
    
    @classmethod
    def quick(
        cls, 
        file_path: Union[str, Path], 
        default_sheet: str = "sheet1"
    ) -> "ExcelManager":
        """
        类方法：快速创建或打开（如果不存在则创建）
        
        Args:
            file_path: 文件路径
            default_sheet: 默认工作表名称
            
        Returns:
            ExcelManager 实例
        """
        return cls(file_path, default_sheet)


# ============================================================================
# 兼容旧版: eExcel (ExcelManager 的别名)
# ============================================================================

eExcel = ExcelManager


# ============================================================================
# ExcelHandler - 面向已有文件的读取/写入工具
# ============================================================================

class ExcelHandler:
    """
    ExcelHandler：面向已有文件的读取/写入工具
    
    这是 ExcelManager 的兼容性包装类，提供类似的接口。
    推荐使用 ExcelManager 类。
    
    特性：
    - 支持自动创建缺失的工作表
    - 提供快速写入（fast_write）以简化范围计算
    
    示例:
        >>> eh = ExcelHandler("data.xlsx")
        >>> eh.write("Sheet1", [[1, 2], [3, 4]], 1, 1, 2, 2)
        >>> data = eh.read("Sheet1", 1, 1, 2, 2)
    """
    
    def __init__(self, file_name: Union[str, Path]):
        """
        初始化 ExcelHandler
        
        Args:
            file_name: Excel 文件路径
        """
        self._manager = ExcelManager(file_name)
    
    def _ensure_sheet(self, sheet_name: str) -> Worksheet:
        """确保工作表存在"""
        return self._manager._ensure_sheet(sheet_name)
    
    def excel_write(
        self, 
        sheet_name: str, 
        results: Sequence[Sequence], 
        start_row: int, 
        start_col: int, 
        end_row: int, 
        end_col: int
    ) -> None:
        """
        写入数据到指定范围
        
        Args:
            sheet_name: 工作表名称
            results: 二维数据
            start_row: 起始行号
            start_col: 起始列号
            end_row: 结束行号
            end_col: 结束列号
        """
        self._manager.write_sheet(
            sheet_name, results, start_row, start_col, 
            end_row, end_col, apply_styles=False
        )
    
    def excel_read(
        self, 
        sheet_name: str, 
        start_row: int, 
        start_col: int, 
        end_row: int, 
        end_col: int
    ) -> List[List[Any]]:
        """
        读取指定范围的数据
        
        Args:
            sheet_name: 工作表名称
            start_row: 起始行号
            start_col: 起始列号
            end_row: 结束行号
            end_col: 结束列号
            
        Returns:
            二维数据列表
        """
        return self._manager.read_sheet(sheet_name, start_row, start_col, end_row, end_col)
    
    def excel_save_as(self, file_name2: Optional[Union[str, Path]] = None) -> None:
        """
        另存为
        
        Args:
            file_name2: 目标文件路径（可选）
        """
        self._manager.save(file_name2)
    
    def excel_quit(self) -> None:
        """关闭并释放资源"""
        self._manager.close()
    
    @staticmethod
    def fast_write(
        sheet_name: str, 
        results: Sequence[Sequence], 
        start_row: int, 
        start_col: int, 
        end_row: int = 0, 
        end_col: int = 0, 
        re: int = 0, 
        xl_book: Optional["ExcelHandler"] = None
    ) -> None:
        """
        静态方法：快速写入（简化范围计算）
        
        Args:
            sheet_name: 工作表名称
            results: 二维数据
            start_row: 起始行号
            start_col: 起始列号
            end_row: 结束行号（re=1 时使用）
            end_col: 结束列号（re=1 时使用）
            re: 0=自动计算, 1=使用显式范围
            xl_book: ExcelHandler 实例
        """
        if xl_book is None:
            raise ValueError("必须提供 xl_book 参数")
        
        actual_end_row, actual_end_col = _auto_range(
            start_row, start_col, results, re, end_row, end_col
        )
        xl_book.excel_write(
            sheet_name, results, start_row, start_col, 
            actual_end_row, actual_end_col
        )


# ============================================================================
# OpenExcel - 通过 Excel 应用打开工作簿
# ============================================================================

class OpenExcel:
    """
    OpenExcel：通过 Excel 应用打开工作簿，适合需要 RefreshAll 的场景
    
    需要安装 Microsoft Excel 才能使用此类。
    
    功能：
    - my_open 上下文：返回 ExcelManager 对象，退出时保存
    - open_save_Excel 上下文：返回 xlwings 的 Workbook，退出时刷新数据并保存
    - file_show：列出工作表并按关键词过滤
    
    示例:
        >>> # 使用上下文管理器自动保存
        >>> with OpenExcel("data.xlsx").my_open() as wb:
        ...     wb.fast_write("Sheet1", [[1, 2]], 1, 1)
        
        >>> # 刷新数据连接（需要 Excel 应用）
        >>> with OpenExcel("data.xlsx").open_save_Excel() as wb:
        ...     wb.api.RefreshAll()
    """
    
    def __init__(self, openfile: Union[str, Path], savefile: Optional[Union[str, Path]] = None):
        """
        初始化 OpenExcel
        
        Args:
            openfile: 要打开的源文件路径
            savefile: 保存目标文件路径（可选，默认为 openfile）
        """
        self.openfile = Path(openfile)
        self.savefile = Path(savefile) if savefile else self.openfile
    
    @contextmanager
    def my_open(self) -> Iterator[ExcelManager]:
        """
        上下文管理器：打开文件并使用 ExcelManager 操作，退出时自动保存
        
        Yields:
            ExcelManager 实例
        """
        manager = None
        try:
            manager = ExcelManager(self.openfile)
            yield manager
            manager.save(self.savefile)
        except Exception as e:
            raise RuntimeError(f"操作 Excel 文件失败: {e}") from e
        finally:
            if manager:
                manager.close()
    
    @contextmanager
    def open_save_Excel(self) -> Iterator[Any]:
        """
        上下文管理器：通过 Excel 应用打开，支持 RefreshAll，退出时保存
        
        注意：此方法需要安装 Microsoft Excel
        
        Yields:
            xlwings Workbook 对象
        """
        app = None
        wb = None
        
        try:
            app = xw.App(visible=False)
            wb = app.books.open(self.openfile)
        except Exception as e:
            if app:
                app.quit()
            raise RuntimeError(f"无法打开 Excel 应用: {e}") from e
        
        try:
            yield wb
        finally:
            try:
                # 刷新所有数据连接
                wb.api.RefreshAll()
                wb.save(self.savefile)
            except Exception as e:
                print(f"警告: 刷新或保存失败: {e}")
            finally:
                if app:
                    app.quit()
    
    def file_show(self, filter: Optional[Union[str, Sequence[str]]] = None) -> List[str]:
        """
        列出工作表并按关键词过滤
        
        Args:
            filter: 过滤关键词（字符串或字符串列表）
            
        Returns:
            工作表名称列表
        """
        app = None
        try:
            app = xw.App(visible=False)
            wb = app.books.open(self.openfile)
            sheet_names = wb.sheet_names
        finally:
            if app:
                app.quit()
        
        if filter is not None:
            filters = [filter] if isinstance(filter, str) else list(filter)
            sheet_names = StringBaba(sheet_names).filter_string_list(filters)
        
        return sheet_names


# ============================================================================
# ExcelOperation - 数据处理类
# ============================================================================

class ExcelOperation:
    """
    ExcelOperation：Excel 数据处理类
    
    功能：
    - split_table：按工作表拆分为多个文件
    - merge_tables：合并多个 Excel 文件
    - convert_to_csv：转换为 CSV 格式
    
    示例:
        >>> # 拆分工作表
        >>> op = ExcelOperation("data.xlsx", "output_folder")
        >>> op.split_table()
        
        >>> # 合并文件
        >>> op.merge_tables(["file1.xlsx", "file2.xlsx"], "merged.xlsx")
    """
    
    def __init__(self, input_file: Union[str, Path], output_folder: Union[str, Path]):
        """
        初始化 ExcelOperation
        
        Args:
            input_file: 输入文件路径
            output_folder: 输出目录路径
        """
        self.input_file = Path(input_file)
        self.output_folder = Path(output_folder)
    
    def split_table(self, sheet_names: Optional[List[str]] = None) -> List[Path]:
        """
        按工作表拆分为多个文件
        
        Args:
            sheet_names: 要拆分的工作表列表（可选，默认全部）
            
        Returns:
            生成的文件路径列表
        """
        if not self.input_file.exists():
            raise FileNotFoundError(f"输入文件不存在: {self.input_file}")
        
        # 创建输出目录
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # 读取 Excel 文件
        excel_file = pd.ExcelFile(self.input_file)
        
        sheets_to_process = sheet_names or excel_file.sheet_names
        generated_files = []
        
        for sheet_name in sheets_to_process:
            if sheet_name not in excel_file.sheet_names:
                print(f"警告: 工作表 '{sheet_name}' 不存在，已跳过")
                continue
            
            try:
                df = pd.read_excel(self.input_file, sheet_name=sheet_name)
                output_file = self.output_folder / f"{sheet_name}.xlsx"
                df.to_excel(output_file, index=False, engine='openpyxl')
                generated_files.append(output_file)
            except Exception as e:
                print(f"警告: 拆分工作表 '{sheet_name}' 失败: {e}")
        
        return generated_files
    
    def merge_tables(
        self, 
        input_files: List[Union[str, Path]], 
        output_file: Union[str, Path],
        sheet_name: str = "Merged"
    ) -> Path:
        """
        合并多个 Excel 文件到一个工作表
        
        Args:
            input_files: 输入文件列表
            output_file: 输出文件路径
            sheet_name: 目标工作表名称
            
        Returns:
            输出文件路径
        """
        all_data = []
        
        for file_path in input_files:
            path = Path(file_path)
            if not path.exists():
                print(f"警告: 文件不存在: {path}")
                continue
            
            try:
                df = pd.read_excel(path)
                all_data.append(df)
            except Exception as e:
                print(f"警告: 读取文件失败 {path}: {e}")
        
        if not all_data:
            raise ValueError("没有有效的数据可以合并")
        
        # 合并数据
        merged_df = pd.concat(all_data, ignore_index=True)
        
        # 保存
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        merged_df.to_excel(output_path, sheet_name=sheet_name, index=False, engine='openpyxl')
        
        return output_path
    
    def convert_to_csv(
        self, 
        sheet_name: Optional[str] = None,
        encoding: str = "utf-8-sig"
    ) -> Path:
        """
        将 Excel 转换为 CSV 格式
        
        Args:
            sheet_name: 工作表名称（可选，默认第一个）
            encoding: 文件编码
            
        Returns:
            CSV 文件路径
        """
        if not self.input_file.exists():
            raise FileNotFoundError(f"输入文件不存在: {self.input_file}")
        
        # 读取数据
        df = pd.read_excel(self.input_file, sheet_name=sheet_name or 0)
        
        # 生成输出路径
        output_file = self.output_folder / f"{self.input_file.stem}.csv"
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # 保存为 CSV
        df.to_csv(output_file, index=False, encoding=encoding)
        
        return output_file


# ============================================================================
# 便捷函数
# ============================================================================

def quick_excel(
    file_path: Union[str, Path], 
    data: Optional[Sequence[Sequence]] = None,
    sheet_name: str = "sheet1",
    start_row: int = 1,
    start_col: int = 1
) -> ExcelManager:
    """
    快速创建 Excel 文件并写入数据（一行代码完成）
    
    Args:
        file_path: 文件路径
        data: 要写入的二维数据（可选）
        sheet_name: 工作表名称
        start_row: 起始行号
        start_col: 起始列号
        
    Returns:
        ExcelManager 实例
        
    示例:
        >>> # 创建空文件
        >>> wb = quick_excel("data.xlsx")
        
        >>> # 创建并写入数据
        >>> wb = quick_excel("data.xlsx", [["Name", "Age"], ["Alice", 25]])
    """
    manager = ExcelManager.create(file_path, sheet_name)
    
    if data:
        manager.fast_write(sheet_name, data, start_row, start_col)
        manager.save()
    
    return manager


def read_excel_quick(
    file_path: Union[str, Path], 
    sheet_name: str = "sheet1",
    as_dataframe: bool = False
) -> Union[List[List[Any]], pd.DataFrame]:
    """
    快速读取 Excel 文件（一行代码完成）
    
    Args:
        file_path: 文件路径
        sheet_name: 工作表名称
        as_dataframe: 是否返回 DataFrame
        
    Returns:
        数据列表或 DataFrame
        
    示例:
        >>> # 读取为列表
        >>> data = read_excel_quick("data.xlsx")
        
        >>> # 读取为 DataFrame
        >>> df = read_excel_quick("data.xlsx", as_dataframe=True)
    """
    with ExcelManager(file_path) as manager:
        if as_dataframe:
            return manager.read_dataframe(sheet_name)
        return manager.read_sheet(sheet_name)


# ============================================================================
# 模块导出
# ============================================================================

__all__ = [
    # 核心类
    "ExcelManager",
    "eExcel",  # 兼容性别名
    "ExcelHandler",
    "OpenExcel",
    "ExcelOperation",
    
    # 便捷函数
    "quick_excel",
    "read_excel_quick",
    "create_workbook",
]
