# !/usr/bin/python
# -*- coding:utf-8 -*-
"""
████─█──█─████─███─█──█─███─███
█──█─█──█─█──█─█───██─█──█──█──
████─████─█──█─███─█─██──█──███
█────█──█─█──█─█───█──█──█──█──
█────█──█─████─███─█──█─███─███
╔╗╔╗╔╗╔═══╗╔══╗╔╗──╔══╗╔══╗╔══╗╔═══╗╔══╗
║║║║║║║╔══╝╚╗╔╝║║──╚╗╔╝║╔╗║║╔╗║╚═╗─║╚╗╔╝
║║║║║║║╚══╗─║║─║║───║║─║╚╝║║║║║─╔╝╔╝─║║─
║║║║║║║╔══╝─║║─║║───║║─║╔╗║║║║║╔╝╔╝──║║─
║╚╝╚╝║║╚══╗╔╝╚╗║╚═╗╔╝╚╗║║║║║╚╝║║─╚═╗╔╝╚╗
╚═╝╚═╝╚═══╝╚══╝╚══╝╚══╝╚╝╚╝╚══╝╚═══╝╚══╝
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

佛祖保佑       永不宕机     永无BUG

~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
@project:home
@author:Phoenix,weiliaozi
@file:pywork
@ide:PyCharm
@date:2023/12/3
@time:17:33
@month:十二月
@email:thisluckyboy@126.com
"""
import base64
import datetime
import os
import pathlib
import re
import shutil
import smtplib
import time
from contextlib import contextmanager
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

import mysql.connector
import openpyxl
import pandas as pd
import pymysql
from openpyxl import load_workbook


def fn_timer(func):
    @wraps(func)
    def function_timer(*args, **kwargs):
        t0 = time.perf_counter()
        result = func(*args, **kwargs)
        t1 = time.perf_counter()
        elapsed_time = t1 - t0
        print(f"Total time running {func.__name__}: {elapsed_time:.2f} seconds")
        return result, elapsed_time

    return function_timer

def decrypt(bs):
    try:
        decoded_bytes = base64.b64decode(bs)
        decoded_str = decoded_bytes.decode("utf-8")
        x = int(decoded_str[6]) + int(decoded_str[-1]) * 10
        # Use list comprehension for building the result list
        result = [decoded_str[i] for i in range(0, len(decoded_str), x)]
        result_str = ''.join(result)
        return result_str
    except Exception as e:
        print(f"Error during decryption: {e}")
        return None

class Database:
    def __init__(self, host, port, user, password, db):
        self.connection_state = 0
        self.connection = None
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.db = db
        self.charset = 'utf8'

    def connect(self):
        try:
            if self.connection_state == 0:
                self.connection = pymysql.connect(
                    host=self.host,
                    port=self.port,
                    user=self.user,
                    password=self.password,
                    db=self.db,
                    charset=self.charset,
                    cursorclass=pymysql.cursors.DictCursor
                )
                self.connection_state = 1
            else:
                self.connection_state = 1
        except Exception as e:
            raise Exception(f"Connection failed: {e}")

    def close(self):
        try:
            if self.connection:
                self.connection.cursor().close()
                self.connection.close()
                self.connection_state = 0
        except Exception as e:
            raise Exception(f"Error closing connection: {e}")

    def execute_sql(self, sql, fetch_all=True, df=False, purchases=None, operation_mode="s"):
        try:
            with self.connection.cursor() as cursor:
                if purchases:
                    cursor.executemany(sql, purchases)
                elif operation_mode == "c":
                    # Assuming sql is the stored procedure name
                    cursor.callproc(sql)
                    cursor.connection.commit()
                else:
                    cursor.execute(sql)

                if fetch_all:
                    if df:
                        return pd.DataFrame(cursor.fetchall())
                        cursor.connection.commit()
                    else:
                        return cursor.fetchall()
                        cursor.connection.commit()
                else:
                    self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise Exception(f"Error executing SQL: {e}")

    def fetchall(self, sql):
        self.connect()
        return self.execute_sql(sql)

    def writesql(self, sql):
        self.connect()
        self.execute_sql(sql, fetch_all=False)

    def writesqlmany(self, sql, purchases):
        self.connect()
        self.execute_sql(sql, fetch_all=False, purchases=purchases)

    def callsql(self, sql):
        self.connect()
        self.execute_sql(sql, fetch_all=False,operation_mode = "c")

    def to_df(self, sql):
        self.connect()
        return pd.DataFrame(self.execute_sql(sql))

    @fn_timer
    def __call__(self, sql, purchases=None, operation_mode="s", i=0):
        if i == 0:
            if operation_mode == "s":
                return self.fetchall(sql)
            elif operation_mode == "w":
                self.writesql(sql)
            elif operation_mode == "c":
                self.callsql(sql)
            elif operation_mode == "wm":
                self.writesqlmany(sql, purchases)
        else:
            return self.to_df(sql)

class FileManagement:
    def __init__(self):
        pass

    def add_prefix(self,filename,file_type):
        pattern = r'[\u4e00-\u9fa5]+'
        matches = re.findall(pattern, filename)[0]
        return f"{matches}.{file_type}"
    def copy_files(self,src_dir, dest_dir, target_files, rename=None,file_type="xls"):
        for target_file in target_files:
            source_path = os.path.join(src_dir, target_file)
            destination_file = rename(target_file,file_type) if rename else target_file
            destination_path = os.path.join(dest_dir, destination_file)
            if os.path.exists(source_path):
                shutil.copy(source_path, destination_path)
                print(f"File {target_file} copied from {source_path} to {destination_path}")
            else:
                print(f"Source file {target_file} not found in the latest folder.")

class MySQLDatabase:
    def __init__(self, config):
        self.config = config
        self.connection = None
        self.connect()

    def connect(self):
        try:
            self.connection = mysql.connector.connect(**self.config)
            print("Connected to MySQL database")
        except mysql.connector.Error as err:
            print(f"Error: {err}")

    def close(self):
        if self.connection:
            self.connection.close()
            print("MySQL connection closed")

    def execute_query(self, query, params=None):
        cursor = self.connection.cursor()
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            self.connection.commit()
            print("Query executed successfully")
        except mysql.connector.Error as err:
            print(f"Error: {err}")
        finally:
            cursor.close()

    def fetch_query(self, query, params=None,dictionary=False):
        cursor = self.connection.cursor(dictionary=dictionary)
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            result = cursor.fetchall()
            return result
        except mysql.connector.Error as err:
            print(f"Error: {err}")
        finally:
            cursor.close()

class ExcelHandler:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)

    def excel_write(self, sheet_name, results, start_row, start_col, end_row, end_col):
        try:
            sheet = self.wb[sheet_name]
            for i, row in enumerate(range(start_row, end_row + 1)):
                for j, value in enumerate(range(start_col, end_col + 1)):
                    sheet.cell(row=row, column=value, value=results[i][j])
            print("Results have been written!")
            self.wb.save(self.file_name)
        except Exception as e:
            print(e)

    def excel_read(self, sheet_name, start_row, start_col, end_row, end_col):
        try:
            sheet = self.wb[sheet_name]
            values = [
                [sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
                for row in range(start_row, end_row + 1)
            ]
            print("Results have been read!")
            return values
        except Exception as e:
            print(e)

    def excel_save_as(self, file_name2):
        try:
            self.wb.save(file_name2)
            print("The file has been saved as " + str(file_name2))
        except Exception as e:
            print(e)

    def excel_quit(self):
        try:
            self.wb.close()
        except Exception as e:
            print(e)

    @staticmethod
    def fast_write(sheet_name, results, start_row, start_col, end_row=0, end_col=0, re=0, xl_book=None):
        if re == 0:
            end_row = len(results) + start_row - 1
            end_col = len(results[0]) + start_col - 1
        elif re == 1:
            pass
        xl_book.excel_write(sheet_name, results, start_row=start_row, start_col=start_col, end_row=end_row, end_col=end_col)


class eFormat(object):
    def __init__(self, results):
        self.results = results

    def toTuple(self):
        try:
            results_sql = "(binary('".encode('utf-8')
            for i in range(len(self.results) - 1):
                results_sql = results_sql + (str(self.results[i][0]) + "'),binary('").encode("utf-8")
            results = results_sql + (str(self.results[len(self.results) - 1][0]) + "'))").encode('utf-8')
            return results
        except Exception as e:
            print(e)
            pass


class eSend(object):
    """
    新邮件系统,可群发,群带附件
    """
    def __init__(self,sender=None,receiver=None,username=None,password=None,smtpserver='smtp.126.com'):
        self.sender = sender
        self.receiver = receiver
        self.username = username
        self.password = password
        self.smtpserver = smtpserver

    def send_email(self, subject,e_content, file_paths, file_names):
        try:
            message = MIMEMultipart()
            message['From'] = self.sender  # 发送
            message['To'] = ",".join(self.receiver)  # 收件
            message['Subject'] = Header(subject, 'utf-8')
            message.attach(MIMEText(e_content, 'plain', 'utf-8'))  # 邮件正文

            # 构造附件群
            for file_path,file_name in zip(file_paths,file_names):
                print(file_name,file_path)
                att1 = MIMEText(open(file_path + file_name, 'rb').read(), 'base64', 'utf-8')
                att1["Content-Type"] = 'application/octet-stream'
                att1.add_header('Content-Disposition', 'attachment', filename=('gbk', '', file_name))
                message.attach(att1)

            # 执行
            smtpSsl=smtplib.SMTP_SSL(self.smtpserver)
            smtpSsl.connect(self.smtpserver,465)  # 连接服务器
            smtpSsl.login(self.username, self.password)  # 登录
            smtpSsl.sendmail(self.sender, self.receiver, message.as_string())  # 发送
            smtpSsl.quit()
            print("The email with file_names has been send!")
        except Exception as e:
            print(e)
            pass

class eConstant(object):
    def __init__(self, interval_day,timeclass=1):
        self.interval_day = interval_day
        self.timeclass=timeclass #1日期 2时间戳 3时刻

    def get_timeparameter(self,Format='%Y%m%d'):
        if self.timeclass==1:
            '返回日期'
            realtime = (datetime.date.today() - datetime.timedelta(days=self.interval_day)).strftime(Format)
        elif self.timeclass==2:
            '返回时间戳'
            realtime = time.localtime(time.time())
        elif self.timeclass==3:
            ':return time'
            if Format=='%Y%m%d':
                Format = '%H%M'
            realtime = time.strftime(Format, time.localtime(time.time()))
        elif self.timeclass==4:
            realtime= datetime.datetime.fromtimestamp(int(time.time()))
        else:
            raise TypeError("你输入的参数不合理!")
        return realtime


class eExcel:
    def __init__(self, visible=1, file_name=None):
        self.file_name = file_name
        if not pathlib.Path(file_name).exists():
            self.create_new_excel(file_name)
        self.wb = openpyxl.load_workbook(file_name)
        self.ws = self.wb.active

    @staticmethod
    def create_new_excel(file_name):
        wb = openpyxl.Workbook()
        wb.save(file_name)

    @staticmethod
    def excel_write(ws, results, start_row, start_col, end_row, end_col):
        for i, row in enumerate(range(start_row, end_row + 1)):
            for j, value in enumerate(range(start_col, end_col + 1)):
                ws.cell(row=row, column=value, value=results[i][j])

    def excel_read(self, start_row, start_col, end_row, end_col):
        valueA = [
            [self.ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
            for row in range(start_row, end_row + 1)
        ]
        return valueA

    def excel_save_as(self, file_name2):
        self.wb.save(file_name2)

    @staticmethod
    def fast_write(ws, results, sr, sc, er=0, ec=0, re=0):
        if re == 0:
            er = len(results) + sr - 1
            ec = len(results[0]) + sc - 1
        elif re == 1:
            pass
        eExcel.excel_write(ws, results, start_row=sr, start_col=sc, end_row=er, end_col=ec)

class OpenExcel:
    def __init__(self, openfile, savefile):
        self.openfile = openfile
        self.savefile = savefile

    @contextmanager
    def my_open(self):
        print(f"Opening Excel file: {self.openfile}")
        xlBook = eExcel(file_name=self.openfile)
        yield xlBook.ws
        xlBook.excel_save_as(self.savefile)