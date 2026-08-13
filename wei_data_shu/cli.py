"""Command-line interface for wei-data-shu."""

from __future__ import annotations

import argparse
import sys
from typing import Sequence

from wei_data_shu.utils import generate_password, search_colors


def _ensure_utf8_stdout() -> None:
    reconfigure = getattr(sys.stdout, "reconfigure", None)
    if callable(reconfigure):
        reconfigure(encoding="utf-8")


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="wei-data-shu", description="Utilities for wei-data-shu")
    subparsers = parser.add_subparsers(dest="command")

    password_parser = subparsers.add_parser("password", help="Generate readable passwords")
    password_parser.add_argument("-l", "--length", type=int, default=13, help="Password length")
    password_parser.add_argument("-c", "--count", type=int, default=1, help="Number of passwords to generate")

    colors_parser = subparsers.add_parser("colors", help="View or search color palette")
    colors_parser.add_argument("query", nargs="?", help="Search by hex, English name, or Chinese name")

    date_parser = subparsers.add_parser("date", help="Print a date offset by N days from today")
    date_parser.add_argument("-d", "--days", type=int, default=0, help="Days to subtract from today (default 0)")
    date_parser.add_argument("-f", "--format", default="%Y-%m-%d", help="strftime format (default %%Y-%%m-%%d)")

    excel_parser = subparsers.add_parser("excel", help="Inspect an Excel workbook")
    excel_sub = excel_parser.add_subparsers(dest="excel_command", required=True)
    info_parser = excel_sub.add_parser("info", help="List sheets and row counts of a workbook")
    info_parser.add_argument("file", help="Path to the .xlsx file")

    return parser


def _run_password(args: argparse.Namespace) -> int:
    if args.count <= 0:
        raise ValueError("count must be greater than 0")
    for _ in range(args.count):
        print(generate_password(args.length))
    return 0


def _run_colors(args: argparse.Namespace) -> int:
    results = search_colors(args.query)
    if not results:
        print(f'No colors matched "{args.query}".')
        return 1

    for record in results:
        index = record["index"]
        color_hex = record["hex"]
        name = record["name"]
        name_zh = record["name_zh"]
        print(f"{index:>2}. {color_hex} | {name} | {name_zh}")
    return 0


def _run_date(args: argparse.Namespace) -> int:
    from wei_data_shu.text import DateFormat

    print(DateFormat(interval_day=args.days, timeclass="date").get_timeparameter(Format=args.format))
    return 0


def _run_excel(args: argparse.Namespace) -> int:
    if args.excel_command != "info":
        return 1
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(args.file, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                print(f"{sheet_name}\t{workbook[sheet_name].max_row} 行")
        finally:
            workbook.close()
    except ImportError:
        print("缺少 excel 依赖，请先安装: pip install 'wei-data-shu[excel]'")
        return 1
    return 0


def main(argv: Sequence[str] | None = None) -> int:
    _ensure_utf8_stdout()
    parser = _build_parser()
    args = parser.parse_args(argv)

    if args.command == "password":
        return _run_password(args)
    if args.command == "colors":
        return _run_colors(args)
    if args.command == "date":
        return _run_date(args)
    if args.command == "excel":
        return _run_excel(args)

    parser.print_help()
    return 0


__all__ = ["main"]
